#!/usr/bin/env python3
"""
受注データ（Excelエクスポート）を商品コード×週（月曜始まり）で集計し、
app_master_data.json に「salesHistory.retail」として追記するスクリプト。

使い方:
    python3 build_sales_history.py <受注データ.xlsx> [--years N]

- 入力Excelの想定シート/列: シート「受注データ」、列 A=受注日 B=受注NO C=商品コード
  D=商品名 E=規格 F=数量 G=単位名 H=売単価 I=売上金額
- 「今のマスタ（retailProducts）に存在する商品コード」だけを対象にする。
  これにより、実商品ではない行（値引き・ポイント・空タル代など）や、
  現在は製造対象外・廃盤になった商品コードは自動的に除外される。
- 直近 N 年分（デフォルト2年）のみを保持する（--years で変更可）。
- 週は月曜始まり（アプリの週定義と一致させるため）。
- 出力は app_master_data.json の "salesHistory": {"retail": {"<code>": {"<週の月曜ISO日付>": 数量, ...}}}
  に上書きマージされる（受注データファイル自体はリポジトリにコミットしないこと。
  金額入りの生の受注明細のため）。
"""
import sys
import json
import argparse
from datetime import timedelta
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
MASTER_PATH = HERE / "app_master_data.json"


def monday_of(d):
    return d - timedelta(days=d.weekday())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--years", type=int, default=2)
    ap.add_argument("--sheet", default="受注データ")
    args = ap.parse_args()

    master = json.loads(MASTER_PATH.read_text(encoding="utf-8"))
    retail_codes = set(p["code"] for p in master["retailProducts"])

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    ws = wb[args.sheet]

    # cutoff: N years back from the most recent order date in the file
    max_date = None
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        if d is not None and (max_date is None or d > max_date):
            max_date = d
    if max_date is None:
        print("エラー: 受注日が1件も見つかりませんでした。シート名・列構成を確認してください。")
        sys.exit(1)
    cutoff = max_date.replace(year=max_date.year - args.years)

    agg = {}
    rows_used = 0
    rows_skipped_not_master = 0
    rows_skipped_old = 0
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        code = ws.cell(row=r, column=3).value
        qty = ws.cell(row=r, column=6).value
        if d is None or code is None or qty is None:
            continue
        code = str(code)
        if code not in retail_codes:
            rows_skipped_not_master += 1
            continue
        if d < cutoff:
            rows_skipped_old += 1
            continue
        wk = monday_of(d).strftime("%Y-%m-%d")
        agg.setdefault(code, {})
        agg[code][wk] = agg[code].get(wk, 0) + qty
        rows_used += 1

    master.setdefault("salesHistory", {})["retail"] = agg

    MASTER_PATH.write_text(
        json.dumps(master, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print(f"最新の受注日: {max_date.date()} / 集計対象期間: {cutoff.date()} 以降")
    print(f"集計対象行数: {rows_used}")
    print(f"除外（現マスタに無い商品コード・非商品行）: {rows_skipped_not_master}")
    print(f"除外（対象期間より古い）: {rows_skipped_old}")
    print(f"履歴を持つ商品数: {len(agg)}")
    print(f"app_master_data.json を更新しました。")


if __name__ == "__main__":
    main()
