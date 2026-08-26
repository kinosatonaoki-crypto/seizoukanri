#!/usr/bin/env python3
"""
卸売の受注データ（Excelエクスポート）を商品コード×週（月曜始まり）で集計し、
app_master_data.json に「salesHistory.wholesale」として追記するスクリプト。
（通販向けの build_sales_history.py と同じ考え方の、卸売版です）

使い方:
    python3 build_wholesale_sales_history.py <卸受注データ.xlsx> [--years N]

- 入力Excelの想定シート/列は、通販の受注データと同じ並びを仮定しています：
  シート「受注データ」、列 A=受注日 B=受注NO C=商品コード D=商品名 E=規格
  F=数量 G=単位名 H=売単価 I=売上金額
  ※実際のファイルが届いたら、シート名・列の並びが上記と違っていないか必ず確認し、
    違う場合はこのスクリプトの --sheet や列番号（COL_* の定数）を実データに合わせて
    修正してください。
- 「今のマスタ（wholesaleProducts）に存在する商品コード」だけを対象にする。
  これにより、送料・箱代・サンプルなど実商品ではない行は自動的に除外される。
- 直近 N 年分（デフォルト2年）のみを保持する（--years で変更可）。
- 週は月曜始まり（アプリの週定義と一致させるため）。
- 出力は app_master_data.json の "salesHistory": {"wholesale": {"<code>": {"<週の月曜ISO日付>": 数量, ...}}}
  に上書きマージされる（受注データファイル自体はリポジトリにコミットしないこと。
  金額入りの生の受注明細のため）。

このスクリプトが app_master_data.json に書き込むと、アプリ側は追加のコード変更なしに
自動的に「週次計画に前年同時期の参考値を表示」「前年実績にリセットボタンで自動入力」が
卸売でも働くようになります（通販とまったく同じ仕組みを共有しているため）。
"""
import sys
import json
import argparse
from datetime import timedelta
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
MASTER_PATH = HERE / "app_master_data.json"

# 列番号（1始まり）。実データの並びが違う場合はここを調整してください。
COL_DATE = 1
COL_CODE = 3
COL_QTY = 6


def monday_of(d):
    return d - timedelta(days=d.weekday())


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--years", type=int, default=2)
    ap.add_argument("--sheet", default="受注データ")
    args = ap.parse_args()

    master = json.loads(MASTER_PATH.read_text(encoding="utf-8"))
    wholesale_codes = set(p["code"] for p in master["wholesaleProducts"])

    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    ws = wb[args.sheet]

    max_date = None
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=COL_DATE).value
        if d is not None and (max_date is None or d > max_date):
            max_date = d
    if max_date is None:
        print("エラー: 受注日が1件も見つかりませんでした。シート名・列構成を確認してください。")
        sys.exit(1)
    cutoff = max_date.replace(year=max_date.year - args.years)

    agg = {}
    rows_used = 0
    rows_skipped_not_master = 0
    rows_skipped_old = 0
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=COL_DATE).value
        code = ws.cell(row=r, column=COL_CODE).value
        qty = ws.cell(row=r, column=COL_QTY).value
        if d is None or code is None or qty is None:
            continue
        code = str(code)
        if code not in wholesale_codes:
            rows_skipped_not_master += 1
            continue
        if d < cutoff:
            rows_skipped_old += 1
            continue
        wk = monday_of(d).strftime("%Y-%m-%d")
        agg.setdefault(code, {})
        agg[code][wk] = agg[code].get(wk, 0) + qty
        rows_used += 1

    master.setdefault("salesHistory", {})["wholesale"] = agg

    MASTER_PATH.write_text(
        json.dumps(master, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print(f"最新の受注日: {max_date.date()} / 集計対象期間: {cutoff.date()} 以降")
    print(f"集計対象行数: {rows_used}")
    print(f"除外（現マスタに無い商品コード・非商品行）: {rows_skipped_not_master}")
    print(f"除外（対象期間より古い）: {rows_skipped_old}")
    print(f"履歴を持つ商品数: {len(agg)} / {len(wholesale_codes)}")
    print(f"app_master_data.json を更新しました。")


if __name__ == "__main__":
    main()
